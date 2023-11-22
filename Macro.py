import time
import math
import sys
from PyQt5.QtWidgets import *
from selenium.webdriver import ActionChains
import traceback

from PyQt5 import uic
from selenium.webdriver.common.by import By
from selenium import webdriver
from PyQt5.QtCore import *
from openpyxl import load_workbook  # 파일 불러오기
from openpyxl import Workbook

class EverStart(QThread):
    def __init__(self, driver, Url):
        super().__init__()
        self.running = True
        self.Can = True
        self.ProductId = "0"
        self.driver =driver
        self.Url = Url
    def run(self):
        try:
            self.driver.get(self.Url)
            self.driver.implicitly_wait(10)
            time.sleep(6)
            Product = self.driver.find_element(By.XPATH,'/html/body/div[1]/div/div/section/div/div[1]/div[2]/p[1]').text
            ProductId = Product.split(' ')[0]
            inven = self.driver.find_element(By.XPATH,'//*[@id="app"]/div/div/section/div/div[1]/div[2]/div[1]/button[1]')
            actions = ActionChains(self.driver)
            actions.move_to_element(inven)
            actions.perform()
            inven.click()
            table = self.driver.find_element(By.XPATH,'/html/body/div[1]/div/div/section/div/div[4]/div/div[2]/div/div[3]/table/tbody')
            tableInfor = table.find_elements(By.CLASS_NAME,'el-table__row')
            Naver_text = ""
            Naver_list = []
            Coupang = []
            cafe = []
            for i in range(len(tableInfor)):
                Naver_list += tableInfor[i].text.split('\n')
                print("_________________________")
            for i in Naver_list:
                if i == '--':
                    i = '0'
                elif i == 'Over 50':
                    i = "50"
                Naver_text += i + '\n'
                Coupang.append(i)
                cafe.append(i)
            # 엑셀 작업
            print(Naver_list)
            Coupang.reverse()
            cafe.reverse()
            self.SaveNaver(ProductId, Naver_text)
            self.SaveCafe24(ProductId, Coupang)
            self.SaveCoupang(ProductId, cafe)
        except Exception as e:
            print(traceback.print_exc())

    def resume(self):
        self.running = True

    def pause(self):
        self.running = False

    def SaveNaver(self, Id, text):
        wb1 = load_workbook("0_네이버.xlsx")
        ws1 = wb1.active
        for x in range(1, ws1.max_row + 1):  # 최대 column의 idx를 가져옴
            if Id in str(ws1.cell(x, 2).value):
                ws1.cell(x,13).value = text
                break
        wb1.save("0_네이버.xlsx")
    def SaveCoupang(self, Id, list_):
        wb2 = load_workbook("0_쿠팡.xlsx")
        ws2 = wb2.active
        for x in range(1, ws2.max_row + 1):  # 최대 column의 idx를 가져옴
            if Id in str(ws2.cell(x, 8).value):
                ws2.cell(x, 19).value = list_.pop()
        wb2.save("0_쿠팡.xlsx")

    def SaveCafe24(self, Id, list_):
        wb3 = load_workbook("0_카페24.xlsx")
        ws3 = wb3.active
        for x in range(1, ws3.max_row + 1):  # 최대 column의 idx를 가져옴
            if Id in str(ws3.cell(x, 2).value):
                ws3.cell(x, 10).value = list_.pop()
        wb3.save("0_카페24.xlsx")



class OzStart(QThread):
    def __init__(self, driver, Url):
        super().__init__()
        self.running = True
        self.Can = True
        self.ProductId = "0"
        self.driver =driver
        self.Url = Url
    def run(self):
        try:
            self.driver.get(self.Url)
            self.driver.implicitly_wait(10)
            time.sleep(7)
            text = self.driver.find_element(By.XPATH,'//*[@id="children"]/div[3]/div[1]/div[1]/span').text
            table = self.driver.find_element(By.XPATH,'/html/body/div/div/div[3]/div[3]/div/div/div/div/div/table/tbody')
            tableInfor = table.find_elements(By.CLASS_NAME, 'ant-table-row')
            NaverText = ''
            Coupang = []
            cafe = []
            for i in range(len(tableInfor)):
                tableInforcells = tableInfor[i].find_elements(By.CLASS_NAME,'ant-input')
                for j in range(len(tableInforcells)):
                    NowCell = tableInforcells[j]
                    style_attribute = NowCell.get_attribute("style")
                    border_color = ""
                    styles = style_attribute.split(";")
                    for style in styles:
                        if "border-color" in style:
                            border_color = style.split(":")[1].strip()
                    if border_color == "rgb(0, 128, 0)":
                        NaverText += "50\n"
                        Coupang.append("50")
                        cafe.append("50")
                    elif border_color == "rgb(238, 238, 238)":
                        NaverText += "0\n"
                        Coupang.append("0")
                        cafe.append("0")
                    else:
                        NaverText += "10\n"
                        Coupang.append("10")
                        cafe.append("10")
                    # print(border_color)
            self.SaveNaver(text, NaverText)
            Coupang.reverse()
            cafe.reverse()
            self.SaveCoupang(text, Coupang)
            self.SaveCafe24(text, cafe)
        except Exception as e:
            print(e)
    def SaveNaver(self, Id, text):
        wb1 = load_workbook("0_네이버.xlsx")
        ws1 = wb1.active
        for x in range(1, ws1.max_row + 1):  # 최대 column의 idx를 가져옴
            if ws1.cell(x, 2).value == Id:
                ws1.cell(x,13).value = text
                break
        wb1.save("0_네이버.xlsx")
    def SaveCoupang(self, Id, list_):
        wb2 = load_workbook("0_쿠팡.xlsx")
        ws2 = wb2.active
        for x in range(1, ws2.max_row + 1):  # 최대 column의 idx를 가져옴
            if Id in str(ws2.cell(x, 8).value):
                ws2.cell(x, 19).value = list_.pop()
        wb2.save("0_쿠팡.xlsx")

    def SaveCafe24(self, Id, list_):
        wb3 = load_workbook("0_카페24.xlsx")
        ws3 = wb3.active
        for x in range(1, ws3.max_row + 1):  # 최대 column의 idx를 가져옴
            if Id in str(ws3.cell(x, 2).value):
                ws3.cell(x, 10).value = list_.pop()
        wb3.save("0_카페24.xlsx")

    def resume(self):
        self.running = True

    def pause(self):
        self.running = False

