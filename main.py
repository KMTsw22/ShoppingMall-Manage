import sys
from PyQt5.QtWidgets import *
from openpyxl.reader.excel import load_workbook
from selenium.webdriver import ActionChains
from datetime import datetime
from PyQt5 import uic
from selenium.webdriver.common.by import By
from selenium import webdriver
from twocaptcha import TwoCaptcha
from Login import EverLogin
from Login import OzLogin
from Macro import EverStart
from Macro import OzStart
from PyQt5.QtCore import *
# from Ui import Ui_Dialog
app = QApplication(sys.argv)

form_class = uic.loadUiType("MainUi.ui")[0]


class MyWindow(QMainWindow, form_class):
    progress_start = pyqtSignal(int)
    progress_finish = pyqtSignal()

    def __init__(self):
        super().__init__()
        self.selen = None
        self.setupUi(self)

        self.EverLoginBtn.clicked.connect(self.LoginEver)
        self.OzLoginBtn.clicked.connect(self.LoginOz)
        self.StartBtn.clicked.connect(self.CrollStart)
        self.driver = webdriver.Chrome()



    def LoginEver(self):
        self.selen = EverLogin(self.driver)
        self.selen.start()
        self.selen.wait()

    def CrollStart(self):
        import time
        try:
            wb = load_workbook("1_EVERUGG_LIST.xlsx")
            ws = wb.active
            EverUrlList = []
            for x in range(1, ws.max_row + 1):  # 최대 column의 idx를 가져옴
                EverUrlList.append(ws.cell(x,2).value)

            for i in range(len(EverUrlList)):
                self.Ever = EverStart(self.driver, EverUrlList[i])
                self.Ever.start()
                self.Ever.wait()
                time.sleep(0.3)
            wb1 = load_workbook("1_OZWEAR_LIST.xlsx")
            ws1 = wb1.active
            OzUrlList = []
            for x in range(1, ws1.max_row + 1):  # 최대 column의 idx를 가져옴
                OzUrlList.append(ws1.cell(x, 2).value)

            for i in range(len(OzUrlList)):
                self.Oz = OzStart(self.driver, OzUrlList[i])
                self.Oz.start()
                self.Oz.wait()
                time.sleep(0.3)
            else:
                wb_naver = load_workbook('0_네이버.xlsx')
                wb_coupang = load_workbook('0_쿠팡.xlsx')
                wb_cafe24 = load_workbook('0_카페24.xlsx')
                ws_naver = wb_naver.active
                ws_coupang = wb_coupang.active
                ws_cafe24 = wb_cafe24.active
                now_time = str(datetime.now())
                time = now_time[:4] + '.' + now_time[5:7] + '.' + now_time[8:10] + '_' + now_time[11:13] + '.' + now_time[14:16]
                wb_naver.save(f"네이버_{time}.xlsx")
                wb_coupang.save(f"쿠팡_{time}.xlsx")
                wb_cafe24.save(f"카페24_{time}.xlsx")
                #파일 초기화
            wb1 = load_workbook("0_네이버.xlsx")
            ws1 = wb1.active
            for x in range(5, ws1.max_row + 1):  # 최대 column의 idx를 가져옴
                ws1.cell(row=x, column=13).value = ""
            wb1.save("0_네이버.xlsx")

            wb2 = load_workbook("0_쿠팡.xlsx")
            ws2 = wb2.active
            for x in range(4, ws2.max_row + 1):  # 최대 column의 idx를 가져옴
                ws2.cell(row=x, column=19).value = ""
            wb2.save("0_쿠팡.xlsx")

            wb3 = load_workbook("0_카페24.xlsx")
            ws3 = wb3.active
            for x in range(1, ws3.max_row + 1):  # 최대 column의 idx를 가져옴
                ws3.cell(row=x, column=10).value = ""
            wb3.save("0_카페24.xlsx")

            self.show_alert("작업이 모두 완료 되었습니다!")
        except Exception as e:
            print(e)
            self.show_alert("작업중 오류가 발생했습니다. 확인후 다시 진행 해주세요")
    # def EverStart(self):
    #     self.selen_1 = EverStart(self.driver)
    #     self.selen_1.start()

    def LoginOz(self):
        self.selen2 = OzLogin(self.driver)
        self.selen2.start()
        self.selen2.wait()

    # def OzStart(self):
    #     self.selen_2 = OzStart(self.driver)
    #     self.selen_2.start()
    #     self.selen_2.wait()


    def pause(self):
        self.selen.pause()

    def show_alert(self, text):
        alert = QMessageBox()
        alert.setWindowTitle("알림")
        alert.setText(text)
        alert.setIcon(QMessageBox.Information)
        alert.setStandardButtons(QMessageBox.Ok)
        alert.exec_()



if __name__ == "__main__":
    myWindow = MyWindow()
    myWindow.show()
    sys.exit(app.exec_())

# python -m PyQt5.uic.pyuic -x LoginUi.ui -o Ui.py
# python -m PyInstaller --onefile --noconsole main.py
